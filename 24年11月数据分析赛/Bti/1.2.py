import pdfplumber
import os
import re
from openpyxl import load_workbook

def extract_product_category(pdf_path):
    try:
        # 使用pdfplumber打开PDF文件
        with pdfplumber.open(pdf_path) as pdf:
            full_text = ""
            # 从每一页提取文本
            for page in pdf.pages:
                full_text += page.extract_text() or ""

        # 使用正则表达式查找“【产品类别】”和“【”之间的内容
        # match = re.search(r"【产品类别】(.*?)【", full_text, re.S)
        # match = re.search(r"【组织状态】(.*?)【", full_text, re.S)
        match = re.search(r"【适用人群】(.*?)【", full_text, re.S)
        if match:
            product_category = match.group(1).strip()
            # 去除空格和句号
            product_category = re.sub(r"[ 。]", "", product_category)
            product_category = re.sub(r"适用于", "", product_category)
            product_category = re.sub(r"[-~]", "～", product_category)
            product_category = re.sub(r"状态", "状况", product_category)
            product_category = re.sub(r"等需要|等需|需要进行|需要", "需", product_category)
            product_category = re.sub(r"及", "和", product_category)
            product_category = re.sub(r"0～12月龄食物蛋白过敏婴儿", "食物蛋白过敏婴儿", product_category)
            product_category = re.sub(r"营养补充", "补充营养", product_category)
            product_category = re.sub(r"岁进", "岁因进", product_category)
            product_category = re.sub(r"\n+", "", product_category)
            return product_category
        return None
    except Exception as e:
        print(f"读取 {pdf_path} 时出错: {e}")  # 处理错误
        return None


def update_excel_with_category(xlsx_path, folder_path):
    workbook = load_workbook(xlsx_path)
    sheet = workbook.active

    # 通过第一行的表头识别列
    headers = {cell.value: idx for idx, cell in enumerate(sheet[1])}

    # 获取“注册证号”和“产品类别”列的索引
    reg_col_idx = headers.get('注册证号')
    # product_col_idx = headers.get('产品类别')
    # product_col_idx = headers.get('组织状态')
    product_col_idx = headers.get('适用人群')

    # 确保找到所需的列
    if reg_col_idx is None or product_col_idx is None:
        print("无法找到所需的列")
        return

    # 遍历指定文件夹中的所有PDF文件
    for file_name in os.listdir(folder_path):
        if file_name.endswith('.pdf'):
            pdf_path = os.path.join(folder_path, file_name)
            # 从PDF提取产品类别
            product_category = extract_product_category(pdf_path)

            if product_category:
                # 使用文件名(去掉扩展名)作为注册证号
                registration_number = os.path.splitext(file_name)[0]

                # 更新Excel中对应的行
                for row in sheet.iter_rows(min_row=2):  # 跳过表头行
                    if row[reg_col_idx].value == registration_number:
                        row[product_col_idx].value = product_category or ""
                        break

    # 保存更改到Excel文件
    workbook.save(xlsx_path)


# 定义PDF文件夹和Excel文件的路径
pdf_folder_path = '泰迪B/数据/特医食品说明书'
excel_file_path = '泰迪B/数据/data1.2对照检查.xlsx'

# 执行更新函数
update_excel_with_category(excel_file_path, pdf_folder_path)
